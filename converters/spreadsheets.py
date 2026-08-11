import asyncio
import logging
import csv
import json
import openpyxl
import xlrd
from fpdf import FPDF
import urllib.request
import os

logger = logging.getLogger(__name__)


def get_dejavu_font():
    font_path = "/tmp/DejaVuSans.ttf"
    if not os.path.exists(font_path):
        try:
            url = "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans.ttf"
            urllib.request.urlretrieve(url, font_path)
        except Exception as e:
            logger.error(f"Failed to download DejaVu font: {e}")
    return font_path


def _xlsx_to_csv(input_path: str, output_path: str) -> str:
    try:
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        return output_path
    except Exception as e:
        logger.error(f"Error converting xlsx to csv: {e}")
        raise


async def convert_xlsx_to_csv(input_path: str, output_path: str) -> str:
    return await asyncio.to_thread(_xlsx_to_csv, input_path, output_path)


def _csv_to_xlsx(input_path: str, output_path: str) -> str:
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(input_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
        wb.save(output_path)
        return output_path
    except Exception as e:
        logger.error(f"Error converting csv to xlsx: {e}")
        raise


async def convert_csv_to_xlsx(input_path: str, output_path: str) -> str:
    return await asyncio.to_thread(_csv_to_xlsx, input_path, output_path)


def _xlsx_to_json(input_path: str, output_path: str) -> str:
    try:
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [str(h) for h in rows[0]]
            data = [dict(zip(headers, row)) for row in rows[1:]]
        else:
            data = []
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return output_path
    except Exception as e:
        logger.error(f"Error converting xlsx to json: {e}")
        raise


async def convert_xlsx_to_json(input_path: str, output_path: str) -> str:
    return await asyncio.to_thread(_xlsx_to_json, input_path, output_path)


def _json_to_csv(input_path: str, output_path: str) -> str:
    try:
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, list) and data:
            headers = list(data[0].keys())
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data)
        return output_path
    except Exception as e:
        logger.error(f"Error converting json to csv: {e}")
        raise


async def convert_json_to_csv(input_path: str, output_path: str) -> str:
    return await asyncio.to_thread(_json_to_csv, input_path, output_path)


def _csv_to_json(input_path: str, output_path: str) -> str:
    try:
        with open(input_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            data = list(reader)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return output_path
    except Exception as e:
        logger.error(f"Error converting csv to json: {e}")
        raise


async def convert_csv_to_json(input_path: str, output_path: str) -> str:
    return await asyncio.to_thread(_csv_to_json, input_path, output_path)


def _xls_to_xlsx(input_path: str, output_path: str) -> str:
    try:
        book = xlrd.open_workbook(input_path)
        sheet = book.sheet_by_index(0)
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(sheet.nrows):
            ws.append(sheet.row_values(r))
        wb.save(output_path)
        return output_path
    except Exception as e:
        logger.error(f"Error converting xls to xlsx: {e}")
        raise


async def convert_xls_to_xlsx(input_path: str, output_path: str) -> str:
    return await asyncio.to_thread(_xls_to_xlsx, input_path, output_path)


def _xlsx_to_pdf(input_path: str, output_path: str) -> str:
    try:
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active

        pdf = FPDF()
        pdf.add_page()
        font_path = get_dejavu_font()
        if os.path.exists(font_path):
            pdf.add_font("DejaVu", "", font_path, uni=True)
            pdf.set_font("DejaVu", size=10)
        else:
            pdf.set_font("Arial", size=10)

        text_lines = []
        for row in ws.iter_rows(values_only=True):
            text_lines.append("\t".join([str(v) if v is not None else "" for v in row]))
        text = "\n".join(text_lines)
        pdf.multi_cell(0, 5, text)
        pdf.output(output_path)
        return output_path
    except Exception as e:
        logger.error(f"Error converting xlsx to pdf: {e}")
        raise


async def convert_xlsx_to_pdf(input_path: str, output_path: str) -> str:
    return await asyncio.to_thread(_xlsx_to_pdf, input_path, output_path)


CONVERTERS = {
    ('xlsx', 'csv'): convert_xlsx_to_csv,
    ('csv', 'xlsx'): convert_csv_to_xlsx,
    ('xlsx', 'json'): convert_xlsx_to_json,
    ('json', 'csv'): convert_json_to_csv,
    ('csv', 'json'): convert_csv_to_json,
    ('xls', 'xlsx'): convert_xls_to_xlsx,
    ('xlsx', 'pdf'): convert_xlsx_to_pdf,
}
